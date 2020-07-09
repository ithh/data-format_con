# from readUT18tir import ReadUT18tir
from pandas import DataFrame, read_csv, set_option
from numpy import sign, exp, sin, arctan, array, float, mean, append, sqrt, cos, tan, log, tanh, cosh
import matplotlib.pyplot as plt
from scipy.optimize import leastsq
import tkinter as tk
from tkinter import filedialog
import time
from openpyxl import load_workbook
from pandas.io.json import json_normalize
import pandas as pd
import json


# TestData = load_workbook('KCT_215.json', read_only=True)  # 根据选择的文件路径读取xlsx格式的实验数据，开启只读模式加快运行速度
BeginTime = time.time()
data_str = open(r'C:\Users\Admin\Desktop\黄辉\KCT_215.json').read()
data_list = json.loads(data_str)

set_option('precision', 20)            # 设置DataFrame精度


# Ifo = TestData['Ifo']  # 实验基本信息存储到Ifo中
# Ifo_index = ['Brand', 'Size', 'Pressure', 'LI']
# Ifo_column = ['Information']
# Information = DataFrame(Ifo.values, index=Ifo_index, columns=Ifo_column)  # 将实验基本信息存入DataFrame中

Index1 = ['alpha', 'kappa', 'camber', 'fz', 'fx', 'fy', 'mx', 'mz', 'pr', 'vx', 'time']
# 准稳态和稳态的数据标签
Index2 = ['fz', 'ver_dis']  # vertical stiff数据标签
Index3 = ['fz', 'long_dis', 'fx']  # long stiff数据标签
Index4 = ['fz', 'lat_dis', 'fy']  # lat stiff数据标签
Index5 = ['fz', 'hor_dis', 'fy']  # Transient_StepSlip数据标签
Index6 = ['Vr', 'Re', 'Fz']  # 有效滚动半径数据标签

# data = [[d['QuasiSteady'],d['score'],d['quote'],d['comment_num']] for d in data_list]

QuasiSteady_SR = data_list["QuasiSteady"]["PureSideSlip"]["Data"]
QS_SR = DataFrame(QuasiSteady_SR, columns=Index1)  # 准稳态纯纵滑实验数据

QuasiSteady_SA = data_list["QuasiSteady"]["PureLongSlip"]["Data"]
QS_SA = DataFrame(QuasiSteady_SA, columns=Index1)  # 准稳态纯侧偏实验数据

QuasiSteady_COM = data_list["QuasiSteady"]["CombinedSlip"]["Data"]
QS_COM = DataFrame(QuasiSteady_COM, columns=Index1)  # 准稳态复合工况实验数据

# Steady_SR = TestData['Steady_SR']
# S_SR = DataFrame(Steady_SR.values, columns=Index1)  # 稳态纯纵滑实验数据
# Steady_SA = TestData['Steady_SA']
# S_SA = DataFrame(Steady_SA.values, columns=Index1)  # 稳态纯侧偏实验数据
# Steady_COM = TestData['Steady_COM']
# S_COM = DataFrame(Steady_COM.values, columns=Index1)  # 稳态复合工况实验数据
Static_Ver = data_list["Static"]["VerStiff"]["Data"]
St_Ver = DataFrame(Static_Ver, columns=Index2)  # 垂向刚度实验数据
Static_Long = data_list["Static"]["LongStiff"]["Data"]
St_long = DataFrame(Static_Long, columns=Index3)  # 纵向刚度实验数据
# Static_Lat = TestData['Static_Lat']
# St_lat = DataFrame(Static_Lat.values, columns=Index4)  # 侧向刚度实验数据
Transient_StepSlip = data_list["Transient"]["StepSlip"]["Data"]
Ts_StepSlip = DataFrame(Transient_StepSlip, columns=Index5)  # 角阶跃实验数据
# ERR = TestData['ERR']
# ERR_Data = DataFrame(ERR.values, columns=Index6)  # 有效滚动半径数据

print('QS_SR',QS_SR)
print('QS_SA',QS_SA)
print('QS_SA',QS_COM)
print('Ts_StepSlip',Ts_StepSlip)

EndTime = time.time()
cost = EndTime - BeginTime
print('本次辨识花费时间' + str(cost) + 's')




def re_xlx():
    BeginTime = time.time()
    """读取表格"""
    TestData = load_workbook(r'C:\Users\Admin\Desktop\黄辉\LINGLONG_215_60R17.xlsx', read_only=True)

    Index1 = ['alpha', 'kappa', 'camber', 'fz', 'fx', 'fy', 'mx', 'mz', 'pr', 'vx', 'time', 'Test Information']
    # 准稳态和稳态的数据标签
    Index2 = ['fz', 'ver_dis', 'Test Information']  # vertical stiff数据标签
    Index3 = ['fz', 'long_dis', 'fx', 'Test Information']  # long stiff数据标签
    Index4 = ['fz', 'lat_dis', 'fy', 'Test Information']  # lat stiff数据标签
    Index5 = ['fz', 'hor_dis', 'fy', 'Test Information']  # Transient_StepSlip数据标签
    Index6 = ['Vr', 'Re', 'Fz']  # 有效滚动半径数据标签

    QuasiSteady_SR = TestData['QuasiSteady_SR']
    QS_SR = DataFrame(QuasiSteady_SR.values, columns=Index1)  # 准稳态纯纵滑实验数据
    QuasiSteady_SA = TestData['QuasiSteady_SA']
    QS_SA = DataFrame(QuasiSteady_SA.values, columns=Index1)  # 准稳态纯侧偏实验数据
    QuasiSteady_COM = TestData['QuasiSteady_COM']
    QS_COM = DataFrame(QuasiSteady_COM.values, columns=Index1)  # 准稳态复合工况实验数据
    # Steady_SR = TestData['Steady_SR']
    # S_SR = DataFrame(Steady_SR.values, columns=Index1)  # 稳态纯纵滑实验数据
    # Steady_SA = TestData['Steady_SA']
    # S_SA = DataFrame(Steady_SA.values, columns=Index1)  # 稳态纯侧偏实验数据
    # Steady_COM = TestData['Steady_COM']
    # S_COM = DataFrame(Steady_COM.values, columns=Index1)  # 稳态复合工况实验数据
    Static_Ver = TestData['Static_Ver']
    St_Ver = DataFrame(Static_Ver.values, columns=Index2)  # 垂向刚度实验数据
    Static_Long = TestData['Static_Long']
    St_long = DataFrame(Static_Long.values, columns=Index3)  # 纵向刚度实验数据
    # Static_Lat = TestData['Static_Lat']
    # St_lat = DataFrame(Static_Lat.values, columns=Index4)  # 侧向刚度实验数据
    Transient_StepSlip = TestData['Transient_StepSlip']
    Ts_StepSlip = DataFrame(Transient_StepSlip.values, columns=Index5)  # 角阶跃实验数据
    # ERR = TestData['ERR']
    # ERR_Data = DataFrame(ERR.values, columns=Index6)  # 有效滚动半径数据


    print('QS_SR',QS_SR)
    EndTime = time.time()
    cost = EndTime - BeginTime
    print('本次辨识花费时间' + str(cost) + 's')
